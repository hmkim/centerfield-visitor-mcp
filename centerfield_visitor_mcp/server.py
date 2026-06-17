"""Centerfield Visitor Reservation MCP Server.

Exposes visitor registration as MCP tools for AI agents.
Supports: single registration, bulk from file (Excel/CSV), bulk from text input.

Usage:
    uvx centerfield-visitor-mcp
"""

import csv
import io
import logging
from pathlib import Path

from mcp.server import FastMCP

from .models import VisitorIn
from .service import register_bulk, register_single

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

mcp = FastMCP("Centerfield Visitor Reservation")


COLUMN_ALIASES = {
    "visitor_name": ["이름", "name", "visitor_name", "성명", "방문자명", "방문자이름"],
    "visitor_company_name": ["회사", "company", "visitor_company_name", "소속", "회사명", "소속회사"],
    "visitor_mobile": ["전화번호", "mobile", "visitor_mobile", "휴대폰", "연락처", "phone", "핸드폰"],
    "visitor_email": ["이메일", "email", "visitor_email", "메일", "e-mail"],
    "visit_date": ["방문일", "date", "visit_date", "날짜", "방문날짜", "방문일자"],
    "visit_time": ["방문시간", "time", "visit_time", "시간"],
    "visit_purpose": ["목적", "purpose", "visit_purpose", "방문목적"],
    "floor": ["층", "floor", "방문층", "층수"],
}


def _normalize_column(col: str) -> str | None:
    col_lower = col.strip().lower()
    for field, aliases in COLUMN_ALIASES.items():
        if col_lower in [a.lower() for a in aliases]:
            return field
    return None


def _map_columns(headers: list[str]) -> dict[int, str]:
    mapping = {}
    for idx, header in enumerate(headers):
        field = _normalize_column(header)
        if field:
            mapping[idx] = field
    return mapping


def _parse_csv_text(text: str) -> list[dict]:
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []

    col_map = _map_columns(rows[0])
    if not col_map:
        return []

    records = []
    for row in rows[1:]:
        if not any(cell.strip() for cell in row):
            continue
        record = {}
        for idx, field in col_map.items():
            if idx < len(row):
                record[field] = row[idx].strip()
        if record.get("visitor_name"):
            records.append(record)
    return records


def _parse_excel(file_path: str) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(cell) if cell else "" for cell in rows[0]]
    col_map = _map_columns(headers)
    if not col_map:
        return []

    records = []
    for row in rows[1:]:
        if not any(cell for cell in row):
            continue
        record = {}
        for idx, field in col_map.items():
            if idx < len(row) and row[idx] is not None:
                value = row[idx]
                if field == "visit_date" and hasattr(value, "strftime"):
                    value = value.strftime("%Y-%m-%d")
                record[field] = str(value).strip()
        if record.get("visitor_name"):
            records.append(record)
    return records


def _build_visitors(records: list[dict]) -> tuple[list[VisitorIn], list[str]]:
    visitors = []
    errors = []
    for i, rec in enumerate(records):
        try:
            rec.setdefault("visit_purpose", "meeting")
            rec.setdefault("floor", "12")
            visitors.append(VisitorIn(**rec))
        except Exception as e:
            errors.append(f"  행 {i + 2}: {e}")

    if errors:
        logger.warning(f"Validation errors:\n" + "\n".join(errors))

    return visitors, errors


@mcp.tool(description="센터필드 빌딩 방문자 1명을 예약 등록합니다. Keywords: 센터필드, 방문자 등록, 방문 예약, 방문 신청, visitor registration, centerfield reservation, register visitor, book visitor")
async def register_visitor(
    visitor_name: str,
    visitor_company_name: str,
    visitor_mobile: str,
    visitor_email: str,
    visit_date: str,
    visit_time: str,
    visit_purpose: str = "meeting",
    floor: str = "12",
) -> str:
    """Register a single visitor to Centerfield building.

    Args:
        visitor_name: 방문자 이름
        visitor_company_name: 방문자 소속 회사명
        visitor_mobile: 방문자 휴대폰 번호 (예: 01012345678)
        visitor_email: 방문자 이메일 주소
        visit_date: 방문 날짜 (YYYY-MM-DD 형식)
        visit_time: 방문 시간 (HH:MM, 30분 단위, 08:00~20:00)
        visit_purpose: 방문 목적 (meeting, visit_business, interview, tour, construction, others)
        floor: 방문 층수 (12 또는 18)
    """
    try:
        visitor = VisitorIn(
            visitor_name=visitor_name,
            visitor_company_name=visitor_company_name,
            visitor_mobile=visitor_mobile,
            visitor_email=visitor_email,
            visit_date=visit_date,
            visit_time=visit_time,
            visit_purpose=visit_purpose,
            floor=floor,
        )
    except Exception as e:
        return f"입력값 오류: {e}"

    result = await register_single(visitor)
    if result.success:
        return f"예약 완료: {visitor_name} ({visit_date} {visit_time})"
    else:
        return f"예약 실패: {visitor_name} - {result.message}"


@mcp.tool(description="Excel 또는 CSV 파일에서 방문자 목록을 읽어 일괄 등록합니다. Keywords: 센터필드, 방문자 일괄 등록, 대량 등록, bulk registration, 엑셀 등록, CSV 등록, 파일로 등록, batch visitor, 방문자 명단")
async def register_visitors_from_file(file_path: str) -> str:
    """Register multiple visitors from an Excel (.xlsx) or CSV file.

    The file must have a header row. Column names can be in Korean or English.
    Recognized columns: 이름/name, 회사/company, 전화번호/mobile, 이메일/email,
    방문일/date, 방문시간/time, 목적/purpose(optional), 층/floor(optional).

    Args:
        file_path: 파일의 절대 경로 (.xlsx 또는 .csv)
    """
    path = Path(file_path)
    if not path.exists():
        return f"파일을 찾을 수 없습니다: {file_path}"

    ext = path.suffix.lower()
    if ext == ".xlsx":
        records = _parse_excel(file_path)
    elif ext == ".csv":
        records = _parse_csv_text(path.read_text(encoding="utf-8-sig"))
    else:
        return f"지원하지 않는 파일 형식입니다: {ext} (xlsx 또는 csv만 지원)"

    if not records:
        return (
            "파일에서 방문자 정보를 추출할 수 없습니다.\n"
            "헤더 행에 다음 컬럼명이 필요합니다: 이름, 회사, 전화번호, 이메일, 방문일, 방문시간\n"
            "파일 내용을 텍스트로 복사하여 register_visitors_from_text 도구를 사용해주세요."
        )

    visitors, errors = _build_visitors(records)

    if not visitors:
        return "유효한 방문자 정보가 없습니다.\n검증 오류:\n" + "\n".join(errors)

    result = await register_bulk(visitors)

    summary = f"일괄 등록 결과: 총 {result.total}명 중 {result.succeeded}명 성공, {result.failed}명 실패"
    if errors:
        summary += f"\n\n파싱 단계에서 건너뛴 행 ({len(errors)}건):\n" + "\n".join(errors)
    if result.failed > 0:
        failed_details = [
            f"  - {r.visitor_name}: {r.message}"
            for r in result.results
            if not r.success
        ]
        summary += "\n\n실패 상세:\n" + "\n".join(failed_details)

    return summary


@mcp.tool(description="텍스트(복사/붙여넣기)로 방문자 목록을 입력받아 일괄 등록합니다. 파일 파싱 실패 시 사용하세요. Keywords: 센터필드, 방문자 등록, 텍스트 입력, 복사 붙여넣기, paste visitors, text registration, 수동 입력, 방문자 목록")
async def register_visitors_from_text(text: str) -> str:
    """Register multiple visitors from pasted text (CSV-like or tab-separated).

    Use this when a file cannot be parsed, or the user copies/pastes visitor data directly.
    The first line must be a header row. Data can be comma-separated or tab-separated.

    Example input:
        이름,회사,전화번호,이메일,방문일,방문시간
        홍길동,ABC주식회사,01012345678,hong@abc.com,2025-01-15,10:00
        김철수,XYZ코퍼레이션,01087654321,kim@xyz.com,2025-01-15,10:30

    Args:
        text: 헤더 포함 방문자 목록 텍스트 (CSV 또는 탭 구분)
    """
    if not text.strip():
        return "입력 텍스트가 비어있습니다. 헤더행과 데이터를 포함해주세요."

    if "\t" in text.split("\n")[0]:
        text = text.replace("\t", ",")

    records = _parse_csv_text(text)

    if not records:
        return (
            "텍스트에서 방문자 정보를 추출할 수 없습니다.\n"
            "첫 줄에 헤더가 필요합니다. 예시:\n"
            "이름,회사,전화번호,이메일,방문일,방문시간\n"
            "홍길동,ABC회사,01012345678,hong@abc.com,2025-01-15,10:00"
        )

    visitors, errors = _build_visitors(records)

    if not visitors:
        return "유효한 방문자 정보가 없습니다.\n검증 오류:\n" + "\n".join(errors)

    result = await register_bulk(visitors)

    summary = f"일괄 등록 결과: 총 {result.total}명 중 {result.succeeded}명 성공, {result.failed}명 실패"
    if errors:
        summary += f"\n\n파싱 단계에서 건너뛴 행 ({len(errors)}건):\n" + "\n".join(errors)
    if result.failed > 0:
        failed_details = [
            f"  - {r.visitor_name}: {r.message}"
            for r in result.results
            if not r.success
        ]
        summary += "\n\n실패 상세:\n" + "\n".join(failed_details)

    return summary


@mcp.tool(description="파일의 방문자 목록을 미리 확인합니다 (실제 등록하지 않음). Keywords: 센터필드, 방문자 미리보기, 파일 확인, preview visitors, 명단 확인, 등록 전 확인")
async def preview_visitors_from_file(file_path: str) -> str:
    """Preview parsed visitor data from a file without registering.
    Use this to verify the data looks correct before submitting.

    Args:
        file_path: 파일의 절대 경로 (.xlsx 또는 .csv)
    """
    path = Path(file_path)
    if not path.exists():
        return f"파일을 찾을 수 없습니다: {file_path}"

    ext = path.suffix.lower()
    if ext == ".xlsx":
        records = _parse_excel(file_path)
    elif ext == ".csv":
        records = _parse_csv_text(path.read_text(encoding="utf-8-sig"))
    else:
        return f"지원하지 않는 파일 형식: {ext}"

    if not records:
        return "파일에서 방문자 정보를 추출할 수 없습니다. 컬럼명을 확인해주세요."

    visitors, errors = _build_visitors(records)

    lines = [f"파싱 결과: {len(visitors)}명 유효, {len(errors)}건 오류\n"]
    for i, v in enumerate(visitors, 1):
        lines.append(
            f"  {i}. {v.visitor_name} | {v.visitor_company_name} | "
            f"{v.visitor_mobile} | {v.visitor_email} | "
            f"{v.visit_date} {v.visit_time} | {v.visit_purpose} | {v.floor}층"
        )

    if errors:
        lines.append(f"\n검증 오류 ({len(errors)}건):")
        lines.extend(errors)

    return "\n".join(lines)


def main():
    mcp.run(transport="stdio")


if __name__ == "__main__":
    main()
